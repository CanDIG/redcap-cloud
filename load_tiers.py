"""
Usage:
  load_tiers.py <metadata_type> <json_filepath> <xls_filepath> <output_filepath>

Options:
  -h --help        Show this screen
  -v --version     Version
  <metadata_type>  one of: clinical, pipeline
  <json_filepath>  Path and filename of the json data file
  <xls_filepath>   Path and filename of the project .xls file
  <output_filepath>Path and filename of the json output
"""


from docopt import docopt
import openpyxl as pyxl
import json

def format_fieldname(field):
	field = str(field).lower()
	split_field = field.split(" ")
	for x in range(1,len(split_field)):
		split_field[x] = split_field[x].capitalize()
	return "".join(split_field)

def load_tiers():
	args = docopt(__doc__, version='0.1')
	metadata_type = args['<metadata_type>']
	json_file = args['<json_filepath>']
	xls_file = args['<xls_filepath>']
	output = args['<output_filepath>']

	wb = pyxl.load_workbook(xls_file)
	sheet = wb.active

	with open(json_file) as f:
		data = json.load(f)

	# TODO: pipeline metadata uses a different key...
	if metadata_type == 'clinical':
		dataset = data['metadata']
	elif metadata_type == 'pipeline':
		dataset = data['pipeline_metadata']
	else:
		raise ValueError('invalid metadata type')

	for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
		table = row[0].value[6:]
		field = row[2].value
		field = format_fieldname(field)
		tier = row[5].value

		for entry in dataset:
			if table in entry:
				entry[table][field+'Tier'] = tier

	data['metadata'] = dataset
	with open(output, 'w') as outfile:
		json.dump(data, outfile)

	print('>>> Tiered dataset output to: {}'.format(output))

if __name__ == "__main__":
    load_tiers()